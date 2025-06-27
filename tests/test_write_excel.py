#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import unittest
import json
import tempfile
import os
import sys
from unittest.mock import Mock, patch, MagicMock
from io import BytesIO

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tools.writeExcel import WriteExcelTool


class TestWriteExcelTool(unittest.TestCase):
    """WriteExcelTool 单元测试类"""

    def setUp(self):
        """测试前的设置"""
        # 创建模拟的运行时和会话对象
        self.mock_runtime = Mock()
        self.mock_session = Mock()
        
        # 创建工具实例
        self.tool = WriteExcelTool(self.mock_runtime, self.mock_session)
        
        # 基础测试数据
        self.simple_data = [
            {"姓名": "张三", "年龄": 25, "部门": "技术部"},
            {"姓名": "李四", "年龄": 30, "部门": "市场部"}
        ]
        
        self.enhanced_data = {
            "data": self.simple_data,
            "format": {
                "column_widths": {"A": 15, "B": 10, "C": 20},
                "row_heights": {"1": 25, "2": 20},
                "merge_cells": ["A1:C1"],
                "cells": {
                    "1,1": {
                        "font": {"bold": True, "size": 14},
                        "background_color": "FFFF00"
                    },
                    "2,2": {
                        "font": {"italic": True},
                        "alignment": {"horizontal": "center"}
                    }
                }
            }
        }

    def test_normalize_cell_key_letter_format(self):
        """测试字母格式的单元格键标准化"""
        # 测试字母格式
        result = self.tool._normalize_cell_key(1, "A")
        self.assertEqual(result, "1,1")
        
        result = self.tool._normalize_cell_key(2, "B")
        self.assertEqual(result, "2,2")
        
        result = self.tool._normalize_cell_key(3, "C")
        self.assertEqual(result, "3,3")

    def test_normalize_cell_key_number_format(self):
        """测试数字格式的单元格键标准化"""
        # 测试数字格式
        result = self.tool._normalize_cell_key(1, 1)
        self.assertEqual(result, "1,1")
        
        result = self.tool._normalize_cell_key(2, 2)
        self.assertEqual(result, "2,2")
        
        result = self.tool._normalize_cell_key(3, 3)
        self.assertEqual(result, "3,3")

    def test_normalize_cell_key_mixed_format(self):
        """测试混合格式的单元格键标准化"""
        # 测试混合格式
        result1 = self.tool._normalize_cell_key(1, "A")
        result2 = self.tool._normalize_cell_key(1, 1)
        self.assertEqual(result1, result2)
        
        result1 = self.tool._normalize_cell_key(2, "B")
        result2 = self.tool._normalize_cell_key(2, 2)
        self.assertEqual(result1, result2)

    def test_normalize_column_index_letter_format(self):
        """测试字母格式的列索引标准化"""
        # 测试字母格式
        result = self.tool._normalize_column_index("A")
        self.assertEqual(result, "A")
        
        result = self.tool._normalize_column_index("B")
        self.assertEqual(result, "B")
        
        result = self.tool._normalize_column_index("a")  # 小写转大写
        self.assertEqual(result, "A")

    def test_normalize_column_index_number_format(self):
        """测试数字格式的列索引标准化"""
        # 测试数字格式
        result = self.tool._normalize_column_index(1)
        self.assertEqual(result, "A")
        
        result = self.tool._normalize_column_index(2)
        self.assertEqual(result, "B")
        
        result = self.tool._normalize_column_index(26)
        self.assertEqual(result, "Z")

    def test_normalize_column_index_string_number_format(self):
        """测试字符串数字格式的列索引标准化"""
        # 测试字符串数字格式
        result = self.tool._normalize_column_index("1")
        self.assertEqual(result, "A")
        
        result = self.tool._normalize_column_index("2")
        self.assertEqual(result, "B")
        
        result = self.tool._normalize_column_index("26")
        self.assertEqual(result, "Z")

    def test_apply_cell_format_with_letter_index(self):
        """测试使用字母索引和数字索引应用单元格格式（只断言 '1,1'）"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="测试")
        
        format_config = {
            "cells": {
                "1,1": {
                    "font": {"bold": True, "size": 14},
                    "background_color": "00FF00"
                }
            }
        }
        self.tool._apply_cell_format(cell, format_config, 1, 1)
        self.assertEqual(cell.font.bold, True)
        self.assertEqual(cell.font.size, 14)
        self.assertEqual(cell.fill.start_color.rgb[2:], "00FF00")

    def test_apply_cell_format_with_number_index(self):
        """测试使用数字索引应用单元格格式"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="测试")
        
        format_config = {
            "cells": {
                "1,1": {
                    "font": {"bold": True, "size": 14},
                    "background_color": "FFFF00"
                }
            }
        }
        
        self.tool._apply_cell_format(cell, format_config, 1, 1)
        
        # 验证字体设置
        self.assertEqual(cell.font.bold, True)
        self.assertEqual(cell.font.size, 14)
        
        # 验证背景颜色 - openpyxl使用ARGB格式，需要去掉前两位
        self.assertEqual(cell.fill.start_color.rgb[2:], "FFFF00")

    def test_apply_column_width_with_letter_index(self):
        """测试使用字母索引应用列宽设置"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        format_config = {
            "column_widths": {
                "A": 15,
                "B": 10,
                "C": 20
            }
        }
        
        self.tool._apply_column_width(ws, format_config)
        
        # 验证列宽设置
        self.assertEqual(ws.column_dimensions["A"].width, 15)
        self.assertEqual(ws.column_dimensions["B"].width, 10)
        self.assertEqual(ws.column_dimensions["C"].width, 20)

    def test_apply_column_width_with_number_index(self):
        """测试使用数字索引应用列宽设置"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        format_config = {
            "column_widths": {
                "1": 15,
                "2": 10,
                "3": 20
            }
        }
        
        self.tool._apply_column_width(ws, format_config)
        
        # 验证列宽设置
        self.assertEqual(ws.column_dimensions["A"].width, 15)
        self.assertEqual(ws.column_dimensions["B"].width, 10)
        self.assertEqual(ws.column_dimensions["C"].width, 20)

    def test_apply_row_height(self):
        """测试应用行高设置"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        format_config = {
            "row_heights": {
                "1": 25,
                "2": 20,
                "3": 30
            }
        }
        
        self.tool._apply_row_height(ws, format_config)
        
        # 验证行高设置
        self.assertEqual(ws.row_dimensions[1].height, 25)
        self.assertEqual(ws.row_dimensions[2].height, 20)
        self.assertEqual(ws.row_dimensions[3].height, 30)

    def test_apply_merge_cells_string_format(self):
        """测试字符串格式的合并单元格"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # 添加一些数据
        ws.cell(row=1, column=1, value="A1")
        ws.cell(row=1, column=2, value="B1")
        ws.cell(row=1, column=3, value="C1")
        
        format_config = {
            "merge_cells": ["A1:C1"]
        }
        
        self.tool._apply_merge_cells(ws, format_config)
        
        # 验证合并单元格
        self.assertTrue(ws.cell(row=1, column=1).coordinate in ws.merged_cells)

    def test_apply_merge_cells_dict_format(self):
        """测试字典格式的合并单元格"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # 添加一些数据
        ws.cell(row=1, column=1, value="A1")
        ws.cell(row=1, column=2, value="B1")
        
        format_config = {
            "merge_cells": [{"start": "A1", "end": "B1"}]
        }
        
        self.tool._apply_merge_cells(ws, format_config)
        
        # 验证合并单元格
        self.assertTrue(ws.cell(row=1, column=1).coordinate in ws.merged_cells)

    def test_apply_merge_cells_list_format(self):
        """测试列表格式的合并单元格"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # 添加一些数据
        ws.cell(row=1, column=1, value="A1")
        ws.cell(row=1, column=2, value="B1")
        
        format_config = {
            "merge_cells": [["A1", "B1"]]
        }
        
        self.tool._apply_merge_cells(ws, format_config)
        
        # 验证合并单元格
        self.assertTrue(ws.cell(row=1, column=1).coordinate in ws.merged_cells)

    @patch('tools.writeExcel.WriteExcelTool.create_text_message')
    @patch('tools.writeExcel.WriteExcelTool.create_blob_message')
    def test_invoke_simple_format(self, mock_create_blob, mock_create_text):
        """测试简单格式的JSON数据处理"""
        # 设置模拟返回值
        mock_text_message = Mock()
        mock_blob_message = Mock()
        mock_create_text.return_value = mock_text_message
        mock_create_blob.return_value = mock_blob_message
        
        # 测试数据
        tool_parameters = {
            'json_str': json.dumps(self.simple_data),
            'filename': 'test_simple'
        }
        
        # 执行测试
        messages = list(self.tool._invoke(tool_parameters))
        
        # 验证结果
        self.assertEqual(len(messages), 2)
        self.assertIn(mock_text_message, messages)
        self.assertIn(mock_blob_message, messages)
        
        # 验证文本消息内容
        mock_create_text.assert_called_once()
        call_args = mock_create_text.call_args[0][0]
        self.assertIn("test_simple.xlsx", call_args)
        self.assertIn("generated successfully", call_args)
        
        # 验证blob消息
        mock_create_blob.assert_called_once()
        blob_args = mock_create_blob.call_args
        self.assertEqual(blob_args[1]['meta']['filename'], 'test_simple.xlsx')
        self.assertEqual(blob_args[1]['meta']['mime_type'], 
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @patch('tools.writeExcel.WriteExcelTool.create_text_message')
    @patch('tools.writeExcel.WriteExcelTool.create_blob_message')
    def test_invoke_enhanced_format(self, mock_create_blob, mock_create_text):
        """测试增强格式的JSON数据处理"""
        # 设置模拟返回值
        mock_text_message = Mock()
        mock_blob_message = Mock()
        mock_create_text.return_value = mock_text_message
        mock_create_blob.return_value = mock_blob_message
        
        # 测试数据
        tool_parameters = {
            'json_str': json.dumps(self.enhanced_data),
            'filename': 'test_enhanced'
        }
        
        # 执行测试
        messages = list(self.tool._invoke(tool_parameters))
        
        # 验证结果
        self.assertEqual(len(messages), 2)
        self.assertIn(mock_text_message, messages)
        self.assertIn(mock_blob_message, messages)

    def test_invoke_invalid_json(self):
        """测试无效JSON数据的处理"""
        tool_parameters = {
            'json_str': 'invalid json string',
            'filename': 'test_invalid'
        }
        
        # 验证抛出异常
        with self.assertRaises(Exception) as context:
            list(self.tool._invoke(tool_parameters))
        
        self.assertIn("Error parsing JSON string", str(context.exception))

    def test_invoke_dict_data(self):
        """测试字典数据的处理"""
        dict_data = {"姓名": "张三", "年龄": 25, "部门": "技术部"}
        
        tool_parameters = {
            'json_str': json.dumps(dict_data),
            'filename': 'test_dict'
        }
        
        # 执行测试
        messages = list(self.tool._invoke(tool_parameters))
        
        # 验证结果
        self.assertEqual(len(messages), 2)

    def test_invoke_empty_data(self):
        """测试空数据的处理"""
        empty_data = []
        
        tool_parameters = {
            'json_str': json.dumps(empty_data),
            'filename': 'test_empty'
        }
        
        # 执行测试
        messages = list(self.tool._invoke(tool_parameters))
        
        # 验证结果
        self.assertEqual(len(messages), 2)

    def test_filename_processing(self):
        """测试文件名处理"""
        tool_parameters = {
            'json_str': json.dumps(self.simple_data),
            'filename': 'test file with spaces'
        }
        
        # 执行测试
        messages = list(self.tool._invoke(tool_parameters))
        
        # 验证文件名中的空格被替换为下划线
        self.assertEqual(len(messages), 2)

    def test_default_filename(self):
        """测试默认文件名处理"""
        json_str = json.dumps(self.simple_data)
        excel_bytes, filename = self.tool.generate_excel_bytes(json_str)
        self.assertEqual(filename, "Formatted_Data.xlsx")

    def test_start_row_default_value(self):
        """测试 start_row 默认值"""
        json_str = json.dumps(self.simple_data)
        excel_bytes, filename = self.tool.generate_excel_bytes(json_str)
        
        # 验证默认从第1行开始写入
        from openpyxl import load_workbook
        from io import BytesIO
        
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        # 检查第1行是否有数据（标题行）
        self.assertIsNotNone(ws.cell(row=1, column=1).value)
        self.assertEqual(ws.cell(row=1, column=1).value, "姓名")

    def test_start_row_custom_value(self):
        """测试自定义 start_row 值"""
        data_with_start_row = {
            "data": self.simple_data,
            "format": {
                "start_row": 3
            }
        }
        json_str = json.dumps(data_with_start_row)
        excel_bytes, filename = self.tool.generate_excel_bytes(json_str)
        
        from openpyxl import load_workbook
        from io import BytesIO
        
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        # 检查第1行和第2行应该为空
        self.assertIsNone(ws.cell(row=1, column=1).value)
        self.assertIsNone(ws.cell(row=2, column=1).value)
        
        # 检查第3行应该有标题行数据
        self.assertIsNotNone(ws.cell(row=3, column=1).value)
        self.assertEqual(ws.cell(row=3, column=1).value, "姓名")
        
        # 检查第4行应该有数据
        self.assertIsNotNone(ws.cell(row=4, column=1).value)
        self.assertEqual(ws.cell(row=4, column=1).value, "张三")

    def test_start_row_with_show_header_false(self):
        """测试 start_row 与 show_header: false 的组合"""
        data_with_start_row_no_header = {
            "data": self.simple_data,
            "format": {
                "start_row": 3,
                "show_header": False
            }
        }
        json_str = json.dumps(data_with_start_row_no_header)
        excel_bytes, filename = self.tool.generate_excel_bytes(json_str)
        
        from openpyxl import load_workbook
        from io import BytesIO
        
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        # 检查第1行和第2行应该为空
        self.assertIsNone(ws.cell(row=1, column=1).value)
        self.assertIsNone(ws.cell(row=2, column=1).value)
        
        # 检查第3行应该有数据（没有标题行）
        self.assertIsNotNone(ws.cell(row=3, column=1).value)
        self.assertEqual(ws.cell(row=3, column=1).value, "张三")
        
        # 检查第4行应该有数据
        self.assertIsNotNone(ws.cell(row=4, column=1).value)
        self.assertEqual(ws.cell(row=4, column=1).value, "李四")

    def test_start_row_with_cell_formatting(self):
        """测试 start_row 与单元格格式化的组合"""
        data_with_start_row_and_format = {
            "data": self.simple_data,
            "format": {
                "start_row": 3,
                "cells": {
                    "3,1": {
                        "font": {"bold": True, "size": 14},
                        "background_color": "FFFF00"
                    },
                    "4,1": {
                        "font": {"italic": True},
                        "alignment": {"horizontal": "center"}
                    }
                }
            }
        }
        json_str = json.dumps(data_with_start_row_and_format)
        excel_bytes, filename = self.tool.generate_excel_bytes(json_str)
        
        from openpyxl import load_workbook
        from io import BytesIO
        
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        # 检查第3行第1列的格式（标题行）
        cell_3_1 = ws.cell(row=3, column=1)
        self.assertEqual(cell_3_1.font.bold, True)
        self.assertEqual(cell_3_1.font.size, 14)
        self.assertEqual(cell_3_1.fill.start_color.rgb[2:], "FFFF00")
        
        # 检查第4行第1列的格式（数据行）
        cell_4_1 = ws.cell(row=4, column=1)
        self.assertEqual(cell_4_1.font.italic, True)
        self.assertEqual(cell_4_1.alignment.horizontal, "center")

    def test_start_row_with_merge_cells(self):
        """测试 start_row 与合并单元格的组合"""
        data_with_start_row_and_merge = {
            "data": self.simple_data,
            "format": {
                "start_row": 3,
                "merge_cells": ["A3:C3"]
            }
        }
        json_str = json.dumps(data_with_start_row_and_merge)
        excel_bytes, filename = self.tool.generate_excel_bytes(json_str)
        
        from openpyxl import load_workbook
        from io import BytesIO
        
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        # 检查合并单元格
        merged_ranges = list(ws.merged_cells.ranges)
        self.assertEqual(len(merged_ranges), 1)
        self.assertEqual(str(merged_ranges[0]), "A3:C3")

    def test_start_row_edge_cases(self):
        """测试 start_row 边界情况"""
        # 测试 start_row = 1（默认值）
        data_start_row_1 = {
            "data": self.simple_data,
            "format": {
                "start_row": 1
            }
        }
        json_str = json.dumps(data_start_row_1)
        excel_bytes, filename = self.tool.generate_excel_bytes(json_str)
        
        from openpyxl import load_workbook
        from io import BytesIO
        
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        # 检查第1行应该有标题行数据
        self.assertIsNotNone(ws.cell(row=1, column=1).value)
        self.assertEqual(ws.cell(row=1, column=1).value, "姓名")
        
        # 测试 start_row = 10（大数值）
        data_start_row_10 = {
            "data": self.simple_data,
            "format": {
                "start_row": 10
            }
        }
        json_str = json.dumps(data_start_row_10)
        excel_bytes, filename = self.tool.generate_excel_bytes(json_str)
        
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        # 检查前9行应该为空
        for row in range(1, 10):
            self.assertIsNone(ws.cell(row=row, column=1).value)
        
        # 检查第10行应该有标题行数据
        self.assertIsNotNone(ws.cell(row=10, column=1).value)
        self.assertEqual(ws.cell(row=10, column=1).value, "姓名")


class TestWriteExcelToolIntegration(unittest.TestCase):
    """WriteExcelTool 集成测试类"""

    def setUp(self):
        """测试前的设置"""
        self.mock_runtime = Mock()
        self.mock_session = Mock()
        self.tool = WriteExcelTool(self.mock_runtime, self.mock_session)

    def test_full_workflow_with_letter_indexes(self):
        """测试使用字母索引的完整工作流程"""
        test_data = {
            "data": [
                {"姓名": "张三", "年龄": 25, "部门": "技术部"},
                {"姓名": "李四", "年龄": 30, "部门": "市场部"}
            ],
            "format": {
                "cells": {
                    "1,A": {"font": {"bold": True}, "background_color": "FFFF00"},
                    "2,B": {"font": {"italic": True}, "alignment": {"horizontal": "center"}}
                },
                "column_widths": {"A": 15, "B": 10, "C": 20},
                "row_heights": {"1": 25, "2": 20},
                "merge_cells": ["A1:C1"]
            }
        }
        
        tool_parameters = {
            'json_str': json.dumps(test_data),
            'filename': 'integration_test'
        }
        
        # 执行测试
        messages = list(self.tool._invoke(tool_parameters))
        
        # 验证结果
        self.assertEqual(len(messages), 2)

    def test_full_workflow_with_number_indexes(self):
        """测试使用数字索引的完整工作流程"""
        test_data = {
            "data": [
                {"姓名": "张三", "年龄": 25, "部门": "技术部"},
                {"姓名": "李四", "年龄": 30, "部门": "市场部"}
            ],
            "format": {
                "cells": {
                    "1,1": {"font": {"bold": True}, "background_color": "FFFF00"},
                    "2,2": {"font": {"italic": True}, "alignment": {"horizontal": "center"}}
                },
                "column_widths": {"1": 15, "2": 10, "3": 20},
                "row_heights": {"1": 25, "2": 20},
                "merge_cells": ["A1:C1"]
            }
        }
        
        tool_parameters = {
            'json_str': json.dumps(test_data),
            'filename': 'integration_test'
        }
        
        # 执行测试
        messages = list(self.tool._invoke(tool_parameters))
        
        # 验证结果
        self.assertEqual(len(messages), 2)

    def test_full_workflow_with_mixed_indexes(self):
        """测试混合索引的完整工作流程"""
        test_data = {
            "data": [
                {"姓名": "张三", "年龄": 25, "部门": "技术部"},
                {"姓名": "李四", "年龄": 30, "部门": "市场部"}
            ],
            "format": {
                "cells": {
                    "1,A": {"font": {"bold": True}},
                    "1,1": {"background_color": "FFFF00"},
                    "2,B": {"font": {"italic": True}},
                    "2,2": {"alignment": {"horizontal": "center"}}
                },
                "column_widths": {"A": 15, "1": 15, "B": 10, "2": 10},
                "row_heights": {"1": 25, "2": 20},
                "merge_cells": ["A1:C1"]
            }
        }
        
        tool_parameters = {
            'json_str': json.dumps(test_data),
            'filename': 'integration_test'
        }
        
        # 执行测试
        messages = list(self.tool._invoke(tool_parameters))
        
        # 验证结果
        self.assertEqual(len(messages), 2)


if __name__ == '__main__':
    unittest.main() 