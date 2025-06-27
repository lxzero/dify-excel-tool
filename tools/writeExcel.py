from collections.abc import Generator
from typing import Any

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage

import pandas as pd
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
import json

class WriteExcelTool(Tool):
    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage]:
        json_str = tool_parameters['json_str']
        filename = tool_parameters.get('filename', 'Formatted Data')
        debug = tool_parameters.get('debug', False)
        excel_bytes, filename_with_ext = self.generate_excel_bytes(json_str, filename)
        if debug:
            with open(filename_with_ext, "wb") as f:
                f.write(excel_bytes)
            yield self.create_text_message(f"[DEBUG] Excel file '{filename_with_ext}' saved to local directory.")
        yield self.create_text_message(f"Excel file '{filename_with_ext}' generated successfully with formatting and merged cells")
        yield self.create_blob_message(
            blob=excel_bytes,
            meta={
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "filename": filename_with_ext
            }
        )

    def generate_excel_bytes(self, jsonData: str, filename: str = "Formatted Data"):
        """生成Excel二进制内容和最终文件名"""
        try:
            data = json.loads(jsonData)
            if isinstance(data, dict) and 'data' in data and 'format' in data:
                df_data = data['data']
                format_config = data.get('format', {})
                if isinstance(df_data, list):
                    df = pd.DataFrame(df_data)
                elif isinstance(df_data, dict):
                    df = pd.DataFrame([df_data])
                else:
                    df = pd.DataFrame(df_data)
            else:
                if isinstance(data, list):
                    df = pd.DataFrame(data)
                elif isinstance(data, dict):
                    df = pd.DataFrame([data])
                else:
                    df = pd.DataFrame(data)
                format_config = {}
        except Exception as e:
            raise Exception(f"Error parsing JSON string: {str(e)}")

        excel_buffer = BytesIO()
        try:
            wb = Workbook()
            ws = wb.active
            show_header = format_config.get('show_header', True)
            # 获取开始行配置，默认为第1行
            start_row = format_config.get('start_row', 1)
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=show_header), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    self._apply_cell_format(cell, format_config, r_idx, c_idx)
            self._apply_column_width(ws, format_config)
            self._apply_row_height(ws, format_config)
            self._apply_merge_cells(ws, format_config)
            wb.save(excel_buffer)
            excel_buffer.seek(0)
        except Exception as e:
            raise Exception(f"Error creating Excel file: {str(e)}")
        filename_with_ext = f"{filename.replace(' ', '_')}.xlsx"
        return excel_buffer.getvalue(), filename_with_ext
    
    def _normalize_cell_key(self, row_idx, col_idx):
        """标准化单元格键，支持字母和数字两种列索引格式"""
        # 如果col_idx是字符串（字母格式），转换为数字
        if isinstance(col_idx, str):
            try:
                col_num = column_index_from_string(col_idx.upper())
                return f"{row_idx},{col_num}"
            except ValueError:
                # 如果转换失败，保持原样
                return f"{row_idx},{col_idx}"
        else:
            # 如果col_idx是数字，直接使用
            return f"{row_idx},{col_idx}"
    
    def _apply_cell_format(self, cell, format_config, row_idx, col_idx):
        """应用单元格格式"""
        # 标准化单元格键，支持字母和数字两种列索引格式
        cell_key = self._normalize_cell_key(row_idx, col_idx)
        
        # 获取单元格特定的格式配置
        cell_format = format_config.get('cells', {}).get(cell_key, {})
        
        # 字体设置
        if 'font' in cell_format:
            font_config = cell_format['font']
            font = Font(
                name=font_config.get('name', 'Calibri'),
                size=font_config.get('size', 11),
                bold=font_config.get('bold', False),
                italic=font_config.get('italic', False),
                color=font_config.get('color', '000000')  # 默认黑色
            )
            cell.font = font
        
        # 背景颜色设置
        if 'background_color' in cell_format:
            fill = PatternFill(
                start_color=cell_format['background_color'],
                end_color=cell_format['background_color'],
                fill_type='solid'
            )
            cell.fill = fill
        
        # 边框设置
        if 'border' in cell_format:
            border_config = cell_format['border']
            border = Border(
                left=Side(style=border_config.get('left', 'thin')),
                right=Side(style=border_config.get('right', 'thin')),
                top=Side(style=border_config.get('top', 'thin')),
                bottom=Side(style=border_config.get('bottom', 'thin'))
            )
            cell.border = border
        
        # 对齐方式设置
        if 'alignment' in cell_format:
            align_config = cell_format['alignment']
            alignment = Alignment(
                horizontal=align_config.get('horizontal', 'left'),
                vertical=align_config.get('vertical', 'bottom'),
                wrap_text=align_config.get('wrap_text', False)
            )
            cell.alignment = alignment
    
    def _normalize_column_index(self, col_idx):
        """标准化列索引，支持字母和数字两种格式"""
        if isinstance(col_idx, str):
            # 如果已经是字母格式，直接返回
            if col_idx.isalpha():
                return col_idx.upper()
            # 如果是数字字符串，转换为字母
            try:
                return get_column_letter(int(col_idx))
            except ValueError:
                return col_idx
        else:
            # 如果是数字，转换为字母
            return get_column_letter(col_idx)
    
    def _apply_column_width(self, ws, format_config):
        """应用列宽设置"""
        column_widths = format_config.get('column_widths', {})
        for col_idx, width in column_widths.items():
            # 标准化列索引
            normalized_col = self._normalize_column_index(col_idx)
            ws.column_dimensions[normalized_col].width = width
    
    def _apply_row_height(self, ws, format_config):
        """应用行高设置"""
        row_heights = format_config.get('row_heights', {})
        for row_num, height in row_heights.items():
            ws.row_dimensions[int(row_num)].height = height
    
    def _apply_merge_cells(self, ws, format_config):
        """应用合并单元格设置"""
        merge_cells = format_config.get('merge_cells', [])
        for merge_range in merge_cells:
            try:
                # 支持多种格式的合并范围
                if isinstance(merge_range, str):
                    # 格式: "A1:B2"
                    ws.merge_cells(merge_range)
                elif isinstance(merge_range, dict):
                    # 格式: {"start": "A1", "end": "B2"}
                    start = merge_range.get('start')
                    end = merge_range.get('end')
                    if start and end:
                        ws.merge_cells(f"{start}:{end}")
                elif isinstance(merge_range, list) and len(merge_range) == 2:
                    # 格式: ["A1", "B2"]
                    start, end = merge_range
                    ws.merge_cells(f"{start}:{end}")
            except Exception as e:
                print(f"Warning: Failed to merge cells {merge_range}: {str(e)}")
                continue